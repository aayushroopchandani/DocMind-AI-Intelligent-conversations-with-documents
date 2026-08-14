from __future__ import annotations

import datetime as dt
import io
import unittest

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side

from scripts.data_analysis_agent.spreadsheet_io import (
    CellType,
    SpreadsheetConversionError,
    SpreadsheetLimits,
    read_xlsx,
    write_xlsx,
)
from scripts.data_analysis_agent.spreadsheet_io.csv_reader import read_csv
from scripts.data_analysis_agent.spreadsheet_io.style_map import resolve_color
from scripts.data_analysis_agent.spreadsheet_io.workbook_model import (
    Cell,
    CellStyle,
    StyleTable,
    WorkbookDocument,
    Worksheet,
)


def _sample_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    for column, heading in enumerate(("Region", "Amount", "Date", "Total"), start=1):
        cell = sheet.cell(row=1, column=column, value=heading)
        cell.font = Font(bold=True, size=12, name="Calibri", color="FFFFFFFF")
        cell.fill = PatternFill(
            fill_type="solid", start_color="FF4472C4", end_color="FF4472C4"
        )
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(bottom=Side(style="medium", color=Color(rgb="FF000000")))
    sheet.append(["North", 1234.5, dt.date(2026, 3, 14), "=B2*2"])
    sheet.append(["South", 990.25, dt.date(2026, 4, 1), "=B3*2"])
    sheet["B2"].number_format = "#,##0.00"
    sheet["C2"].number_format = "yyyy-mm-dd"
    sheet.merge_cells("A5:C5")
    sheet["A5"] = "merged note"
    sheet.freeze_panes = "B2"
    sheet.column_dimensions["A"].width = 18
    sheet.row_dimensions[1].height = 28

    hidden = workbook.create_sheet("Notes")
    hidden["A1"] = "hello"
    hidden.sheet_state = "hidden"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _cell_at(sheet: Worksheet, row: int, column: int) -> Cell:
    return next(
        cell for cell in sheet.cells if cell.row == row and cell.column == column
    )


class XlsxReaderTests(unittest.TestCase):
    def setUp(self) -> None:
        self.document = read_xlsx(_sample_workbook(), name="sample.xlsx")

    def test_reads_sheets_in_order_with_visibility(self) -> None:
        self.assertEqual(
            [(sheet.name, sheet.hidden) for sheet in self.document.sheets],
            [("Sales", False), ("Notes", True)],
        )

    def test_formula_cells_carry_the_formula_not_a_cached_value(self) -> None:
        cell = _cell_at(self.document.sheets[0], 1, 3)
        self.assertIs(cell.type, CellType.FORMULA)
        self.assertEqual(cell.formula, "=B2*2")
        self.assertIsNone(cell.value)

    def test_dates_become_excel_serial_numbers(self) -> None:
        cell = _cell_at(self.document.sheets[0], 1, 2)
        self.assertIs(cell.type, CellType.DATE)
        # 2026-03-14 under the 1900 epoch.
        self.assertEqual(cell.value, 46095.0)

    def test_styles_are_interned_and_resolved(self) -> None:
        header = _cell_at(self.document.sheets[0], 0, 0)
        second = _cell_at(self.document.sheets[0], 0, 1)
        self.assertIsNotNone(header.style_id)
        self.assertEqual(header.style_id, second.style_id)

        style = self.document.styles[header.style_id]
        self.assertTrue(style.bold)
        self.assertEqual(style.text_color, "#FFFFFF")
        self.assertEqual(style.background_color, "#4472C4")
        self.assertTrue(style.wrap_text)
        self.assertIsNotNone(style.border_bottom)

    def test_layout_metadata_survives(self) -> None:
        sheet = self.document.sheets[0]
        self.assertEqual(sheet.merges, ["A5:C5"])
        self.assertEqual(sheet.frozen_rows, 1)
        self.assertEqual(sheet.frozen_columns, 1)
        self.assertEqual(sheet.columns[0].width, 18.0)
        self.assertEqual(sheet.rows[0].height, 28.0)

    def test_rejects_files_that_are_not_workbooks(self) -> None:
        with self.assertRaises(SpreadsheetConversionError) as caught:
            read_xlsx(b"not a workbook at all", name="broken.xlsx")
        self.assertEqual(caught.exception.code, "unreadable_workbook")

    def test_enforces_cell_limits(self) -> None:
        with self.assertRaises(SpreadsheetConversionError) as caught:
            read_xlsx(
                _sample_workbook(),
                name="sample.xlsx",
                limits=SpreadsheetLimits(max_cells_per_sheet=3),
            )
        self.assertEqual(caught.exception.code, "sheet_too_large")


class XlsxRoundTripTests(unittest.TestCase):
    def test_values_formulas_and_layout_survive_a_round_trip(self) -> None:
        first = read_xlsx(_sample_workbook(), name="sample.xlsx")
        second = read_xlsx(write_xlsx(first), name="round.xlsx")

        def fingerprint(document: WorkbookDocument) -> set[tuple[object, ...]]:
            sheet = document.sheets[0]
            return {
                (cell.row, cell.column, cell.type, cell.value, cell.formula)
                for cell in sheet.cells
            }

        self.assertEqual(fingerprint(first), fingerprint(second))
        self.assertEqual(second.sheets[0].merges, ["A5:C5"])
        self.assertEqual(second.sheets[0].frozen_rows, 1)
        self.assertEqual(second.sheets[0].columns[0].width, 18.0)
        self.assertTrue(second.sheets[1].hidden)

    def test_styles_survive_a_round_trip(self) -> None:
        first = read_xlsx(_sample_workbook(), name="sample.xlsx")
        second = read_xlsx(write_xlsx(first), name="round.xlsx")
        before = first.styles[_cell_at(first.sheets[0], 0, 0).style_id]
        after = second.styles[_cell_at(second.sheets[0], 0, 0).style_id]
        self.assertEqual(before, after)

    def test_number_formats_survive_a_round_trip(self) -> None:
        first = read_xlsx(_sample_workbook(), name="sample.xlsx")
        second = read_xlsx(write_xlsx(first), name="round.xlsx")
        amount = _cell_at(second.sheets[0], 1, 1)
        self.assertEqual(second.styles[amount.style_id].number_format, "#,##0.00")

    def test_sheet_titles_are_made_excel_safe_and_unique(self) -> None:
        document = WorkbookDocument(
            name="odd.xlsx",
            sheets=[
                Worksheet(name="a/b:c", index=0, row_count=0, column_count=0),
                Worksheet(name="a/b:c", index=1, row_count=0, column_count=0),
            ],
        )
        result = read_xlsx(write_xlsx(document), name="odd.xlsx")
        self.assertEqual([sheet.name for sheet in result.sheets], ["a_b_c", "a_b_c (2)"])


class StyleMapTests(unittest.TestCase):
    def test_resolves_argb_theme_and_indexed_colours(self) -> None:
        self.assertEqual(resolve_color(Color(rgb="FF4472C4")), "#4472C4")
        # Theme slot 4 is accent 1 in the default Office palette.
        self.assertEqual(resolve_color(Color(theme=4, tint=0.0)), "#4472C4")
        self.assertIsNone(resolve_color(None))

    def test_tint_lightens_and_darkens(self) -> None:
        lighter = resolve_color(Color(theme=1, tint=0.5))
        self.assertIsNotNone(lighter)
        self.assertNotEqual(lighter, "#000000")


class StyleTableTests(unittest.TestCase):
    def test_default_styles_are_not_stored(self) -> None:
        table = StyleTable(limit=10)
        self.assertIsNone(table.intern(CellStyle()))
        self.assertEqual(table.as_list(), [])

    def test_identical_styles_share_one_entry(self) -> None:
        table = StyleTable(limit=10)
        first = table.intern(CellStyle(bold=True))
        second = table.intern(CellStyle(bold=True))
        self.assertEqual(first, second)
        self.assertEqual(len(table.as_list()), 1)

    def test_style_table_is_bounded(self) -> None:
        table = StyleTable(limit=1)
        table.intern(CellStyle(bold=True))
        with self.assertRaises(SpreadsheetConversionError) as caught:
            table.intern(CellStyle(italic=True))
        self.assertEqual(caught.exception.code, "too_many_styles")


class CsvReaderTests(unittest.TestCase):
    def test_reads_values_and_types(self) -> None:
        document = read_csv(b"name,qty\nwidget,4\n", name="items.csv")
        sheet = document.sheets[0]
        self.assertEqual(sheet.row_count, 2)
        self.assertEqual(sheet.column_count, 2)
        self.assertIs(_cell_at(sheet, 1, 1).type, CellType.NUMBER)
        self.assertEqual(_cell_at(sheet, 1, 1).value, 4.0)

    def test_keeps_leading_zero_codes_as_text(self) -> None:
        document = read_csv(b"code\n007\n", name="codes.csv")
        cell = _cell_at(document.sheets[0], 1, 0)
        self.assertIs(cell.type, CellType.STRING)
        self.assertEqual(cell.value, "007")

    def test_detects_semicolon_delimiters(self) -> None:
        document = read_csv(b"a;b;c\n1;2;3\n", name="euro.csv")
        self.assertEqual(document.sheets[0].column_count, 3)

    def test_strips_a_utf8_byte_order_mark(self) -> None:
        document = read_csv("﻿name,qty\n".encode("utf-8"), name="bom.csv")
        self.assertEqual(_cell_at(document.sheets[0], 0, 0).value, "name")


class InterchangeContractTests(unittest.TestCase):
    """The frontend emits this exact JSON shape.

    `extra="forbid"` on the models means any drift between the TypeScript
    types in `lib/data-analysis/sheet/workbook-document.ts` and the Pydantic
    schema fails here rather than in production.
    """

    PAYLOAD: dict[str, object] = {
        "schema_version": 1,
        "name": "Untitled Analysis",
        "date_system": "1900",
        "sheets": [
            {
                "name": "Sheet1",
                "index": 0,
                "hidden": False,
                "tab_color": None,
                "row_count": 3,
                "column_count": 3,
                "cells": [
                    {
                        "row": 0,
                        "column": 0,
                        "value": "Region",
                        "formula": None,
                        "type": "string",
                        "style_id": 0,
                    },
                    {
                        "row": 1,
                        "column": 1,
                        "value": 1234.5,
                        "formula": None,
                        "type": "number",
                        "style_id": 1,
                    },
                    {
                        "row": 1,
                        "column": 2,
                        "value": 46095.0,
                        "formula": None,
                        "type": "date",
                        "style_id": 2,
                    },
                    {
                        "row": 2,
                        "column": 1,
                        "value": None,
                        "formula": "=SUM(B2:B2)",
                        "type": "formula",
                        "style_id": None,
                    },
                    {
                        "row": 2,
                        "column": 2,
                        "value": True,
                        "formula": None,
                        "type": "boolean",
                        "style_id": None,
                    },
                ],
                "merges": ["A4:C4"],
                "columns": [{"index": 0, "width": 18.0, "hidden": False}],
                "rows": [{"index": 0, "height": 21.0, "hidden": False}],
                "frozen_rows": 1,
                "frozen_columns": 0,
                "show_gridlines": True,
            }
        ],
        "styles": [
            {
                "font_family": "Arial",
                "font_size": 12.0,
                "bold": True,
                "italic": False,
                "underline": False,
                "strikethrough": False,
                "text_color": "#FFFFFF",
                "background_color": "#4472C4",
                "horizontal": "center",
                "vertical": "middle",
                "wrap_text": True,
                "number_format": None,
                "border_top": None,
                "border_bottom": {"style": "medium", "color": "#000000"},
                "border_left": None,
                "border_right": None,
            },
            {
                "font_family": None,
                "font_size": None,
                "bold": False,
                "italic": False,
                "underline": False,
                "strikethrough": False,
                "text_color": None,
                "background_color": None,
                "horizontal": None,
                "vertical": None,
                "wrap_text": False,
                "number_format": "#,##0.00",
                "border_top": None,
                "border_bottom": None,
                "border_left": None,
                "border_right": None,
            },
            {
                "font_family": None,
                "font_size": None,
                "bold": False,
                "italic": False,
                "underline": False,
                "strikethrough": False,
                "text_color": None,
                "background_color": None,
                "horizontal": None,
                "vertical": None,
                "wrap_text": False,
                "number_format": "yyyy-mm-dd",
                "border_top": None,
                "border_bottom": None,
                "border_left": None,
                "border_right": None,
            },
        ],
        "warnings": [],
    }

    def test_frontend_payload_parses_and_renders(self) -> None:
        document = WorkbookDocument.model_validate(self.PAYLOAD)
        result = read_xlsx(write_xlsx(document), name="contract.xlsx")

        sheet = result.sheets[0]
        self.assertEqual(sheet.merges, ["A4:C4"])
        self.assertEqual(sheet.frozen_rows, 1)
        self.assertEqual(sheet.columns[0].width, 18.0)

        self.assertIs(_cell_at(sheet, 1, 1).type, CellType.NUMBER)
        self.assertIs(_cell_at(sheet, 1, 2).type, CellType.DATE)
        self.assertEqual(_cell_at(sheet, 2, 1).formula, "=SUM(B2:B2)")
        self.assertIs(_cell_at(sheet, 2, 2).type, CellType.BOOLEAN)

        header = result.styles[_cell_at(sheet, 0, 0).style_id]
        self.assertTrue(header.bold)
        self.assertEqual(header.background_color, "#4472C4")
        self.assertTrue(header.wrap_text)

    def test_unknown_fields_are_rejected(self) -> None:
        payload = {**self.PAYLOAD, "unexpected": True}
        with self.assertRaises(ValueError):
            WorkbookDocument.model_validate(payload)


if __name__ == "__main__":  # pragma: no cover
    unittest.main()
