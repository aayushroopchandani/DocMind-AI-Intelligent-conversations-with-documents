"""XLSX bytes to the interchange model, via openpyxl 3.1.

Two deliberate choices:

* The workbook is opened with ``data_only=False``, so formula cells carry
  their formula rather than Excel's last cached result. The grid recalculates
  on load, which is the honest behaviour — a cached value can be stale or
  simply absent if the file was written by a tool that never calculated.
* Normal (not read-only) mode is used. Read-only mode drops merged ranges and
  depends on the writing application declaring correct dimensions, and the
  size caps in ``limits`` already bound what can be opened.
"""

from __future__ import annotations

import datetime as dt
import io
from typing import Any, Final

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import CALENDAR_MAC_1904, to_excel
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet

from scripts.data_analysis_agent.spreadsheet_io.limits import (
    DEFAULT_LIMITS,
    SpreadsheetConversionError,
    SpreadsheetLimits,
    ensure,
)
from scripts.data_analysis_agent.spreadsheet_io.style_map import read_cell_style
from scripts.data_analysis_agent.spreadsheet_io.workbook_model import (
    Cell,
    CellType,
    ColumnMeta,
    ConversionWarning,
    DateSystem,
    RowMeta,
    StyleTable,
    WorkbookDocument,
    Worksheet,
)


_SECONDS_PER_DAY: Final[float] = 86_400.0

#: Features that exist in the file but cannot cross into the grid. Reported
#: rather than silently dropped, so the user knows what they are not seeing.
_UNSUPPORTED_FEATURES: Final[tuple[tuple[str, str, str], ...]] = (
    ("_charts", "charts_dropped", "chart"),
    ("_images", "images_dropped", "image"),
    ("_pivots", "pivot_tables_dropped", "pivot table"),
)


class _WarningCollector:
    """Accumulates per-sheet warnings, collapsing repeats into counts."""

    __slots__ = ("_warnings",)

    def __init__(self) -> None:
        self._warnings: dict[tuple[str, str | None], ConversionWarning] = {}

    def add(
        self, code: str, message: str, *, sheet: str | None = None, count: int = 1
    ) -> None:
        key = (code, sheet)
        existing = self._warnings.get(key)
        if existing is None:
            self._warnings[key] = ConversionWarning(
                code=code, message=message, sheet=sheet, count=count
            )
            return
        self._warnings[key] = existing.model_copy(
            update={"count": existing.count + count}
        )

    def as_list(self) -> list[ConversionWarning]:
        return list(self._warnings.values())


def _serial_from_datetime(value: Any, epoch: dt.datetime) -> float | None:
    """Excel serial number for a date-like value, or None if unsupported."""

    if isinstance(value, dt.timedelta):
        return value.total_seconds() / _SECONDS_PER_DAY
    if isinstance(value, dt.time) and not isinstance(value, dt.datetime):
        return (
            value.hour * 3600 + value.minute * 60 + value.second
        ) / _SECONDS_PER_DAY + value.microsecond / (_SECONDS_PER_DAY * 1_000_000)
    if isinstance(value, (dt.datetime, dt.date)):
        try:
            return float(to_excel(value, epoch))
        except (ValueError, OverflowError):
            # Dates before the epoch have no serial representation.
            return None
    return None


def _read_cell(
    cell: Any,
    *,
    epoch: dt.datetime,
    styles: StyleTable,
    limits: SpreadsheetLimits,
    warnings: _WarningCollector,
    sheet_name: str,
) -> Cell | None:
    """Convert one openpyxl cell, or None when it holds nothing at all."""

    value = cell.value
    data_type = cell.data_type
    style = read_cell_style(cell, cell.number_format)
    style_id = styles.intern(style)

    if value is None:
        # An empty cell still matters when it carries formatting.
        if style_id is None:
            return None
        return Cell(
            row=cell.row - 1,
            column=cell.column - 1,
            type=CellType.BLANK,
            style_id=style_id,
        )

    formula: str | None = None
    cell_type: CellType
    resolved: str | float | bool | None

    if data_type == "f" or (isinstance(value, str) and value.startswith("=")):
        formula = value if isinstance(value, str) else str(value)
        ensure(
            len(formula) <= limits.max_formula_length,
            "formula_too_long",
            f"A formula in “{sheet_name}” exceeds "
            f"{limits.max_formula_length} characters.",
        )
        cell_type = CellType.FORMULA
        resolved = None
    elif data_type == "e":
        cell_type = CellType.ERROR
        resolved = str(value)
    elif isinstance(value, bool):
        cell_type = CellType.BOOLEAN
        resolved = value
    elif isinstance(value, (dt.datetime, dt.date, dt.time, dt.timedelta)):
        serial = _serial_from_datetime(value, epoch)
        if serial is None:
            warnings.add(
                "date_out_of_range",
                "Dates before 1900 cannot be represented and were kept as text.",
                sheet=sheet_name,
            )
            cell_type = CellType.STRING
            resolved = str(value)
        else:
            cell_type = CellType.DATE
            resolved = serial
    elif isinstance(value, (int, float)):
        cell_type = CellType.NUMBER
        resolved = float(value)
    else:
        text = str(value)
        if len(text) > limits.max_string_length:
            text = text[: limits.max_string_length]
            warnings.add(
                "text_truncated",
                f"Cell text longer than {limits.max_string_length} characters "
                "was truncated.",
                sheet=sheet_name,
            )
        cell_type = CellType.STRING
        resolved = text

    return Cell(
        row=cell.row - 1,
        column=cell.column - 1,
        value=resolved,
        formula=formula,
        type=cell_type,
        style_id=style_id,
    )


def _read_freeze(sheet: OpenpyxlWorksheet) -> tuple[int, int]:
    """Frozen row and column counts from the sheet's freeze anchor."""

    anchor = sheet.freeze_panes
    if not anchor:
        return 0, 0
    try:
        from openpyxl.utils.cell import coordinate_to_tuple

        row, column = coordinate_to_tuple(str(anchor))
    except (ValueError, TypeError):
        return 0, 0
    return max(0, row - 1), max(0, column - 1)


def _read_dimensions(
    sheet: OpenpyxlWorksheet, limits: SpreadsheetLimits
) -> tuple[list[ColumnMeta], list[RowMeta]]:
    columns: list[ColumnMeta] = []
    for dimension in sheet.column_dimensions.values():
        if dimension.width is None and not dimension.hidden:
            continue
        # A single dimension record can span a run of columns.
        start = dimension.min or 1
        end = min(dimension.max or start, limits.max_columns_per_sheet)
        width = (
            float(dimension.width)
            if dimension.width and dimension.customWidth
            else None
        )
        if width is None and not dimension.hidden:
            continue
        for index in range(start, end + 1):
            columns.append(
                ColumnMeta(
                    index=index - 1, width=width, hidden=bool(dimension.hidden)
                )
            )

    rows: list[RowMeta] = []
    for dimension in sheet.row_dimensions.values():
        height = (
            float(dimension.height)
            if dimension.height and dimension.customHeight
            else None
        )
        if height is None and not dimension.hidden:
            continue
        index = int(dimension.index or 1) - 1
        if index < 0 or index >= limits.max_rows_per_sheet:
            continue
        rows.append(
            RowMeta(index=index, height=height, hidden=bool(dimension.hidden))
        )

    columns.sort(key=lambda item: item.index)
    rows.sort(key=lambda item: item.index)
    return columns, rows


def _read_sheet(
    sheet: OpenpyxlWorksheet,
    *,
    index: int,
    epoch: dt.datetime,
    styles: StyleTable,
    limits: SpreadsheetLimits,
    warnings: _WarningCollector,
    remaining_cells: int,
) -> Worksheet:
    name = str(sheet.title)
    max_row = min(int(sheet.max_row or 0), limits.max_rows_per_sheet)
    max_column = min(int(sheet.max_column or 0), limits.max_columns_per_sheet)

    if (sheet.max_row or 0) > limits.max_rows_per_sheet:
        warnings.add(
            "rows_truncated",
            f"Only the first {limits.max_rows_per_sheet} rows were imported.",
            sheet=name,
        )
    if (sheet.max_column or 0) > limits.max_columns_per_sheet:
        warnings.add(
            "columns_truncated",
            f"Only the first {limits.max_columns_per_sheet} columns were "
            "imported.",
            sheet=name,
        )

    cells: list[Cell] = []
    if max_row and max_column:
        for row in sheet.iter_rows(
            min_row=1, max_row=max_row, min_col=1, max_col=max_column
        ):
            for cell in row:
                converted = _read_cell(
                    cell,
                    epoch=epoch,
                    styles=styles,
                    limits=limits,
                    warnings=warnings,
                    sheet_name=name,
                )
                if converted is None:
                    continue
                cells.append(converted)
                ensure(
                    len(cells) <= limits.max_cells_per_sheet,
                    "sheet_too_large",
                    f"Sheet “{name}” has more than "
                    f"{limits.max_cells_per_sheet} populated cells.",
                )
                ensure(
                    len(cells) <= remaining_cells,
                    "workbook_too_large",
                    "Workbook has more than "
                    f"{limits.max_cells_total} populated cells.",
                )

    merges = [str(item) for item in sheet.merged_cells.ranges]
    ensure(
        len(merges) <= limits.max_merges_per_sheet,
        "too_many_merges",
        f"Sheet “{name}” has more than {limits.max_merges_per_sheet} merged "
        "ranges.",
    )

    for attribute, code, label in _UNSUPPORTED_FEATURES:
        items = getattr(sheet, attribute, None) or ()
        if items:
            warnings.add(
                code,
                f"{len(items)} {label}(s) could not be imported.",
                sheet=name,
                count=len(items),
            )
    if getattr(sheet, "conditional_formatting", None) and list(
        sheet.conditional_formatting
    ):
        warnings.add(
            "conditional_formatting_dropped",
            "Conditional formatting rules were not imported.",
            sheet=name,
        )
    if getattr(sheet, "data_validations", None) and sheet.data_validations.dataValidation:
        warnings.add(
            "data_validation_dropped",
            "Data validation rules were not imported.",
            sheet=name,
        )

    frozen_rows, frozen_columns = _read_freeze(sheet)
    columns, rows = _read_dimensions(sheet, limits)
    tab_color = None
    raw_tab_color = getattr(sheet.sheet_properties, "tabColor", None)
    if raw_tab_color is not None:
        from scripts.data_analysis_agent.spreadsheet_io.style_map import (
            resolve_color,
        )

        tab_color = resolve_color(raw_tab_color)

    return Worksheet(
        name=name,
        index=index,
        hidden=str(sheet.sheet_state) != "visible",
        tab_color=tab_color,
        row_count=max_row,
        column_count=max_column,
        cells=cells,
        merges=merges,
        columns=columns,
        rows=rows,
        frozen_rows=frozen_rows,
        frozen_columns=frozen_columns,
        show_gridlines=bool(sheet.sheet_view.showGridLines),
    )


def read_xlsx(
    content: bytes,
    *,
    name: str,
    limits: SpreadsheetLimits = DEFAULT_LIMITS,
) -> WorkbookDocument:
    """Convert XLSX bytes into the interchange model.

    The caller is expected to have validated the bytes already (zip shape,
    macro rejection, size) — this function is about fidelity, not safety.
    """

    warnings = _WarningCollector()
    try:
        workbook: Workbook = load_workbook(
            io.BytesIO(content),
            read_only=False,
            data_only=False,
            keep_vba=False,
            keep_links=False,
            rich_text=False,
        )
    except SpreadsheetConversionError:
        raise
    except Exception as exc:  # openpyxl raises a wide range of parse errors
        raise SpreadsheetConversionError(
            "unreadable_workbook",
            "This file could not be read as an Excel workbook.",
        ) from exc

    try:
        sheet_names = list(workbook.sheetnames)
        ensure(
            len(sheet_names) <= limits.max_sheets,
            "too_many_sheets",
            f"Workbook has more than {limits.max_sheets} sheets.",
        )

        epoch = workbook.epoch
        styles = StyleTable(limits.max_styles)
        sheets: list[Worksheet] = []
        used_cells = 0

        for index, sheet_name in enumerate(sheet_names):
            sheet = workbook[sheet_name]
            converted = _read_sheet(
                sheet,
                index=index,
                epoch=epoch,
                styles=styles,
                limits=limits,
                warnings=warnings,
                remaining_cells=limits.max_cells_total - used_cells,
            )
            used_cells += len(converted.cells)
            sheets.append(converted)

        if workbook.defined_names:
            warnings.add(
                "defined_names_dropped",
                "Named ranges were not imported.",
                count=len(workbook.defined_names),
            )

        return WorkbookDocument(
            name=name,
            date_system=(
                DateSystem.EXCEL_1904
                if epoch == CALENDAR_MAC_1904
                else DateSystem.EXCEL_1900
            ),
            sheets=sheets,
            styles=styles.as_list(),
            warnings=warnings.as_list(),
        )
    finally:
        workbook.close()


def used_range_a1(sheet: Worksheet) -> str:
    """`A1:D20` for a converted sheet — handy for logging and tests."""

    if not sheet.row_count or not sheet.column_count:
        return "A1"
    return f"A1:{get_column_letter(sheet.column_count)}{sheet.row_count}"
