"""The interchange model back to XLSX bytes.

Written in openpyxl's normal mode rather than ``write_only``: the model is
sparse and unordered, and write-only mode can only append whole rows in
sequence. The cell caps in ``limits`` keep the in-memory workbook bounded.

Dates are written as their serial number plus a date number format, which is
exactly how Excel stores them — no datetime round-trip, so no timezone or
1900-leap-year surprises introduced on the way out.
"""

from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Color
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import CALENDAR_MAC_1904
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet

from scripts.data_analysis_agent.spreadsheet_io.limits import (
    DEFAULT_LIMITS,
    SpreadsheetConversionError,
    SpreadsheetLimits,
    ensure,
)
from scripts.data_analysis_agent.spreadsheet_io.style_map import (
    OpenpyxlStyleFactory,
)
from scripts.data_analysis_agent.spreadsheet_io.workbook_model import (
    Cell,
    CellType,
    DateSystem,
    WorkbookDocument,
    Worksheet,
)


#: Applied when a date cell arrives without a format of its own, so the
#: serial number does not surface as a bare five-digit integer.
_FALLBACK_DATE_FORMAT = "yyyy-mm-dd"

_INVALID_TITLE_CHARACTERS = str.maketrans({character: "_" for character in r"[]:*?/\\"})


def _safe_sheet_title(name: str, taken: set[str]) -> str:
    """Excel rejects some characters and caps titles at 31 characters."""

    cleaned = (name or "Sheet").translate(_INVALID_TITLE_CHARACTERS).strip() or "Sheet"
    cleaned = cleaned[:31]
    if cleaned not in taken:
        taken.add(cleaned)
        return cleaned
    for suffix in range(2, 1000):
        marker = f" ({suffix})"
        candidate = f"{cleaned[: 31 - len(marker)]}{marker}"
        if candidate not in taken:
            taken.add(candidate)
            return candidate
    raise SpreadsheetConversionError(
        "duplicate_sheet_name", f"Cannot find a unique name for sheet “{name}”."
    )


def _write_cell(
    sheet: OpenpyxlWorksheet,
    cell: Cell,
    *,
    styles: list[Any],
    factory: OpenpyxlStyleFactory,
) -> None:
    target = sheet.cell(row=cell.row + 1, column=cell.column + 1)

    if cell.type is CellType.FORMULA and cell.formula:
        target.value = cell.formula
    elif cell.type is CellType.BLANK:
        target.value = None
    elif cell.type is CellType.BOOLEAN:
        target.value = bool(cell.value)
    elif cell.type in (CellType.NUMBER, CellType.DATE):
        target.value = cell.value if isinstance(cell.value, (int, float)) else None
    elif cell.value is not None:
        target.value = str(cell.value)

    parts: dict[str, Any] = {}
    if cell.style_id is not None and 0 <= cell.style_id < len(styles):
        parts = factory.build(cell.style_id, styles[cell.style_id])

    for attribute in ("font", "fill", "alignment", "border"):
        value = parts.get(attribute)
        if value is not None:
            setattr(target, attribute, value)

    number_format = parts.get("number_format")
    if number_format:
        target.number_format = number_format
    elif cell.type is CellType.DATE:
        target.number_format = _FALLBACK_DATE_FORMAT


def _write_sheet(
    sheet: OpenpyxlWorksheet,
    source: Worksheet,
    *,
    document: WorkbookDocument,
    limits: SpreadsheetLimits,
    factory: OpenpyxlStyleFactory,
) -> None:
    ensure(
        len(source.cells) <= limits.max_cells_per_sheet,
        "sheet_too_large",
        f"Sheet “{source.name}” has more than "
        f"{limits.max_cells_per_sheet} cells.",
    )

    for cell in source.cells:
        ensure(
            cell.row < limits.max_rows_per_sheet
            and cell.column < limits.max_columns_per_sheet,
            "cell_out_of_range",
            f"Sheet “{source.name}” references a cell outside the supported "
            "grid.",
        )
        _write_cell(sheet, cell, styles=document.styles, factory=factory)

    for merge in source.merges[: limits.max_merges_per_sheet]:
        try:
            sheet.merge_cells(merge)
        except ValueError:
            # A malformed range should not fail the whole export.
            continue

    # openpyxl derives `customWidth`/`customHeight` from the value itself,
    # so setting the size is enough — the flags are read-only properties.
    for column in source.columns:
        dimension = sheet.column_dimensions[get_column_letter(column.index + 1)]
        if column.width is not None:
            dimension.width = column.width
        if column.hidden:
            dimension.hidden = True

    for row in source.rows:
        dimension = sheet.row_dimensions[row.index + 1]
        if row.height is not None:
            dimension.height = row.height
        if row.hidden:
            dimension.hidden = True

    if source.frozen_rows or source.frozen_columns:
        sheet.freeze_panes = (
            f"{get_column_letter(source.frozen_columns + 1)}"
            f"{source.frozen_rows + 1}"
        )

    sheet.sheet_view.showGridLines = source.show_gridlines
    if source.hidden:
        sheet.sheet_state = "hidden"
    if source.tab_color:
        sheet.sheet_properties.tabColor = Color(
            rgb=f"FF{source.tab_color.lstrip('#').upper()}"
        )


def write_xlsx(
    document: WorkbookDocument,
    *,
    limits: SpreadsheetLimits = DEFAULT_LIMITS,
) -> bytes:
    """Render the interchange model as a `.xlsx` file."""

    ensure(
        len(document.sheets) <= limits.max_sheets,
        "too_many_sheets",
        f"Workbook has more than {limits.max_sheets} sheets.",
    )
    ensure(
        document.cell_count <= limits.max_cells_total,
        "workbook_too_large",
        f"Workbook has more than {limits.max_cells_total} cells.",
    )

    workbook = Workbook()
    # A fresh workbook ships with one sheet; sheets are created explicitly.
    workbook.remove(workbook.active)
    if document.date_system is DateSystem.EXCEL_1904:
        workbook.epoch = CALENDAR_MAC_1904

    factory = OpenpyxlStyleFactory()
    taken: set[str] = set()
    ordered = sorted(document.sheets, key=lambda item: item.index)

    for source in ordered or [Worksheet(name="Sheet1", index=0, row_count=0, column_count=0)]:
        sheet = workbook.create_sheet(title=_safe_sheet_title(source.name, taken))
        _write_sheet(
            sheet,
            source,
            document=document,
            limits=limits,
            factory=factory,
        )

    from io import BytesIO

    buffer = BytesIO()
    try:
        workbook.save(buffer)
    except SpreadsheetConversionError:
        raise
    except Exception as exc:  # pragma: no cover - openpyxl serialisation guard
        raise SpreadsheetConversionError(
            "workbook_not_writable",
            "This workbook could not be written as an Excel file.",
        ) from exc
    finally:
        workbook.close()
    return buffer.getvalue()
